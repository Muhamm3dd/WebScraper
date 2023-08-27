import requests
from bs4 import BeautifulSoup
from openpyxl import Workbook
import os

from openpyxl.drawing.image import Image

base_url = 'https://www.ourcommons.ca'
url = "https://www.ourcommons.ca/members/en/search"
req = requests.get(url)
soup = BeautifulSoup(req.content, 'html.parser')
information = []

def get_names(tag):
    if tag.has_attr("class") and tag["class"] == ["ce-mip-mp-name"]:
        return True
    return False

def get_province(tag):
    if tag.has_attr("class") and tag["class"] == ["ce-mip-mp-province"]:
        return True
    return False

def get_constituent(tag):
    if tag.has_attr("class") and tag["class"] == ["ce-mip-mp-constituency"]:
        return True
    return False

def get_political_party(tag):
    if tag.has_attr("class") and tag["class"] == ["ce-mip-mp-party"]:
        return True
    return False

def get_image(tag):
    if tag.has_attr("class") and tag["class"] == ["ce-mip-mp-picture-container"]:
        return True
    return False

def get_page_url(tag):
    if tag.has_attr("class") and tag["class"] == ["ce-mip-mp-tile"]:
        return True
    return False

content = soup.body.find_all(attrs={"class": 'ce-mip-mp-tile-container'})

for tag in content:
    names = tag.find(get_names)
    provinces = tag.find(get_province)
    constituencies = tag.find(get_constituent)
    political_parties = tag.find(get_political_party)
    images = tag.find_all(get_image)
    page_urls = tag.find_all(get_page_url)

    # Extract the URLs from 'img' and 'a' tags
    image_url = None
    if images:
        image_url = images[0].find('img')['src']
        if image_url.startswith('/'):
            image_url = f"{base_url}{image_url}"

    member_page_url = None
    if page_urls:
        # Loop through page_urls to find the correct URL for each member
        for url_tag in page_urls:
            if url_tag.has_attr("href"):
                member_page_url = f"{base_url}{url_tag['href']}"
                break

    for name, province, constituency, party in zip(names, provinces, constituencies, political_parties):
        information.append({
            "name": name.text.strip(),
            "province": province.text.strip(),
            "constituency": constituency.text.strip(),
            "party": party.text.strip(),
            "image_url": image_url,
            "member_page_url": member_page_url
        })

wb = Workbook()
workSheet = wb.active
workSheet.column_dimensions['A'].width = 20  # Set column width for the Image column
workSheet['A1'] = "Names"
workSheet.column_dimensions['B'].width = 20  # Set column width for the Image column
workSheet['B1'] = "Province"
workSheet.column_dimensions['C'].width = 20  # Set column width for the Image column
workSheet['C1'] = "Constituency"
workSheet.column_dimensions['D'].width = 20  # Set column width for the Image column
workSheet['D1'] = "Party"
workSheet.column_dimensions['E'].width = 20  # Set column width for the Image column
workSheet['E1'] = "Image"  # Add Image column after Party
workSheet.column_dimensions['F'].width = 90 # Set column width for the Image column
workSheet['F1'] = "Image URL"
workSheet.column_dimensions['G'].width = 60  # Set column width for the Image column
workSheet['G1'] = "Member URL"

row = 2
for info in information:
    workSheet.cell(row=row, column=1).value = info["name"]
    workSheet.cell(row=row, column=2).value = info["province"]
    workSheet.cell(row=row, column=3).value = info["constituency"]
    workSheet.cell(row=row, column=4).value = info["party"]
    workSheet.cell(row=row, column=6).value = info["image_url"]
    workSheet.cell(row=row, column=7).value = info["member_page_url"]

    # Load and insert the image into the worksheet
    if info["image_url"]:
        if info["image_url"].startswith('/'):
            image_url = f"{base_url}{info['image_url']}"
        else:
            image_url = info["image_url"]

        image_response = requests.get(image_url)
        if image_response.status_code == 200:
            image_path = f"{info['name']}.jpg"  # Unique image path based on the member's name
            with open(image_path, "wb") as f:
                f.write(image_response.content)

            img = Image(image_path)
            img.width = img.height = 120  # Resize the image to fit the cell
            workSheet.column_dimensions['E'].width = 20  # Adjust the column width to fit the image
            workSheet.row_dimensions[row].height = 70  # Set row height to fit the image
            workSheet.add_image(img, f'E{row}')

    row += 1

wb.save("Canadian_Members_Data.xlsx")

ht = img.anchor(90)
if ht = 9
    print("task done ")

elif print("task failed, try again")
